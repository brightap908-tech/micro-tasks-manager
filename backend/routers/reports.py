"""
Reporting and statistics endpoints.
Generates earnings reports, completion stats, and CSV/Excel exports.
"""

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, case
from typing import List
from datetime import datetime, date, timedelta, timezone
import csv
import io
import json

from backend.database import get_db
from backend import models, schemas

router = APIRouter(prefix="/api/reports", tags=["reports"])


def _parse_date(d: str | None) -> datetime | None:
    if d is None:
        return None
    try:
        return datetime.strptime(d, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    except ValueError:
        return None


@router.get("/dashboard", response_model=schemas.DashboardStats)
def get_dashboard_stats(db: Session = Depends(get_db)):
    today_start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    week_start = today_start - timedelta(days=today_start.weekday())

    tasks = db.query(models.Task).all()
    completed = [t for t in tasks if t.status == models.TaskStatus.COMPLETED]
    pending = [t for t in tasks if t.status == models.TaskStatus.PENDING]
    in_progress = [t for t in tasks if t.status == models.TaskStatus.IN_PROGRESS]
    skipped = [t for t in tasks if t.status == models.TaskStatus.SKIPPED]

    today_tasks = [
        t for t in completed
        if t.completed_at and t.completed_at.replace(tzinfo=timezone.utc) >= today_start
    ]
    week_tasks = [
        t for t in completed
        if t.completed_at and t.completed_at.replace(tzinfo=timezone.utc) >= week_start
    ]

    websites = db.query(models.Website).all()

    return schemas.DashboardStats(
        total_earnings=sum(t.reward for t in completed),
        tasks_completed=len(completed),
        tasks_pending=len(pending),
        tasks_in_progress=len(in_progress),
        tasks_skipped=len(skipped),
        connected_websites=len(websites),
        active_websites=sum(1 for w in websites if w.is_enabled),
        time_spent_today_seconds=sum(t.time_spent_seconds for t in today_tasks),
        time_spent_week_seconds=sum(t.time_spent_seconds for t in week_tasks),
    )


@router.get("/earnings/by-website", response_model=List[schemas.EarningsByWebsite])
def earnings_by_website(
    from_date: str = None,
    to_date: str = None,
    db: Session = Depends(get_db)
):
    q = db.query(models.Task).filter(models.Task.status == models.TaskStatus.COMPLETED)
    fd = _parse_date(from_date)
    td = _parse_date(to_date)
    if fd:
        q = q.filter(models.Task.completed_at >= fd)
    if td:
        q = q.filter(models.Task.completed_at <= td)
    tasks = q.all()

    by_website: dict[int, dict] = {}
    for t in tasks:
        wid = t.website_id or 0
        if wid not in by_website:
            w = db.query(models.Website).filter_by(id=wid).first()
            by_website[wid] = {
                "website_id": wid,
                "website_name": w.name if w else "Unknown",
                "total_earnings": 0.0,
                "task_count": 0,
                "completed_count": 0,
            }
        by_website[wid]["total_earnings"] += t.reward
        by_website[wid]["task_count"] += 1
        by_website[wid]["completed_count"] += 1

    return list(by_website.values())


@router.get("/earnings/by-category", response_model=List[schemas.EarningsByCategory])
def earnings_by_category(
    from_date: str = None,
    to_date: str = None,
    db: Session = Depends(get_db)
):
    q = db.query(models.Task)
    fd = _parse_date(from_date)
    td = _parse_date(to_date)
    if fd:
        q = q.filter(models.Task.completed_at >= fd)
    if td:
        q = q.filter(models.Task.completed_at <= td)
    tasks = q.all()

    by_cat: dict[str, dict] = {}
    for t in tasks:
        cat = t.category.value if t.category else "other"
        if cat not in by_cat:
            by_cat[cat] = {"category": cat, "total_earnings": 0.0, "task_count": 0, "completed_count": 0}
        by_cat[cat]["task_count"] += 1
        if t.status == models.TaskStatus.COMPLETED:
            by_cat[cat]["total_earnings"] += t.reward
            by_cat[cat]["completed_count"] += 1

    return list(by_cat.values())


@router.get("/daily", response_model=List[schemas.DailyStats])
def daily_stats(days: int = Query(30, le=365), db: Session = Depends(get_db)):
    result = []
    today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    for i in range(days - 1, -1, -1):
        day_start = today - timedelta(days=i)
        day_end = day_start + timedelta(days=1)
        tasks = db.query(models.Task).filter(
            models.Task.status == models.TaskStatus.COMPLETED,
            models.Task.completed_at >= day_start,
            models.Task.completed_at < day_end,
        ).all()
        result.append(schemas.DailyStats(
            date=day_start.strftime("%Y-%m-%d"),
            earnings=sum(t.reward for t in tasks),
            tasks_completed=len(tasks),
            time_spent_seconds=sum(t.time_spent_seconds for t in tasks),
        ))
    return result


@router.get("/activity", response_model=List[schemas.ActivityLog])
def get_activity(limit: int = Query(50, le=200), db: Session = Depends(get_db)):
    return (
        db.query(models.ActivityLog)
        .order_by(models.ActivityLog.created_at.desc())
        .limit(limit)
        .all()
    )


@router.get("/export/csv")
def export_csv(from_date: str = None, to_date: str = None, db: Session = Depends(get_db)):
    """Export completed tasks as a CSV file."""
    q = db.query(models.Task)
    fd = _parse_date(from_date)
    td = _parse_date(to_date)
    if fd:
        q = q.filter(models.Task.completed_at >= fd)
    if td:
        q = q.filter(models.Task.completed_at <= td)
    tasks = q.order_by(models.Task.completed_at.desc()).all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "Title", "Category", "Status", "Reward", "Currency",
                     "Website", "Time (s)", "Created", "Completed", "Notes"])
    for t in tasks:
        website_name = t.website.name if t.website else ""
        writer.writerow([
            t.id, t.title, t.category.value if t.category else "",
            t.status.value, t.reward, t.currency, website_name,
            t.time_spent_seconds,
            t.created_at.strftime("%Y-%m-%d %H:%M:%S") if t.created_at else "",
            t.completed_at.strftime("%Y-%m-%d %H:%M:%S") if t.completed_at else "",
            t.notes or "",
        ])

    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode()),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=tasks_export.csv"},
    )


@router.get("/export/excel")
def export_excel(from_date: str = None, to_date: str = None, db: Session = Depends(get_db)):
    """Export tasks as an Excel file using openpyxl."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    q = db.query(models.Task)
    fd = _parse_date(from_date)
    td = _parse_date(to_date)
    if fd:
        q = q.filter(models.Task.completed_at >= fd)
    if td:
        q = q.filter(models.Task.completed_at <= td)
    tasks = q.order_by(models.Task.completed_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tasks"

    headers = ["ID", "Title", "Category", "Status", "Reward", "Currency",
               "Website", "Time (min)", "Created", "Completed", "Notes"]
    header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row, t in enumerate(tasks, 2):
        website_name = t.website.name if t.website else ""
        ws.append([
            t.id, t.title, t.category.value if t.category else "",
            t.status.value, t.reward, t.currency, website_name,
            round(t.time_spent_seconds / 60, 2),
            t.created_at.strftime("%Y-%m-%d %H:%M:%S") if t.created_at else "",
            t.completed_at.strftime("%Y-%m-%d %H:%M:%S") if t.completed_at else "",
            t.notes or "",
        ])

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].auto_size = True

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=tasks_export.xlsx"},
    )
